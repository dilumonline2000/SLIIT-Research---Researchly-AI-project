# -*- coding: utf-8 -*-
"""Builds four expert-evaluation validation-report workbooks, one per module,
from live test evidence gathered by running each FastAPI service locally and
inspecting real responses (see _results_module{1..4}.json + server logs)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULT_COLORS = {
    "PASS": "C6EFCE",
    "PARTIAL": "FFEB9C",
    "FAIL": "FFC7CE",
    "BLOCKED": "D9D9D9",
}
RESULT_FONT = {
    "PASS": "006100",
    "PARTIAL": "9C6500",
    "FAIL": "9C0006",
    "BLOCKED": "404040",
}
HEADER_FILL = "4338CA"
SEV_COLORS = {"High": "FFC7CE", "Medium": "FFEB9C", "Low": "E2EFDA", "": "FFFFFF", "N/A": "FFFFFF"}

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
        cell.border = BORDER


def write_summary_sheet(wb, module_no, module_name, owner, service, counts, env_notes, narrative, key_findings):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 90

    r = 1
    ws.cell(row=r, column=1, value=f"Module {module_no} — {module_name}").font = Font(bold=True, size=16, color=HEADER_FILL)
    r += 1
    ws.cell(row=r, column=1, value="Expert Evaluation / Output Validation Report").font = Font(size=12, italic=True, color="6B7280")
    r += 2

    meta = [
        ("Module owner", owner),
        ("Service under test", service),
        ("Validation method", "Live functional testing — each ML microservice was started locally with real trained "
         "model artifacts and a real Supabase/Gemini/Semantic Scholar connection; every test case below was executed "
         "against the running service and the actual response was inspected and graded (not simulated)."),
        ("Validated by", "Automated technical validation (Claude, AI coding assistant) — panel/supervisor sign-off pending"),
        ("Date of validation", "28 August 2026"),
        ("Report scope", "Read/inference endpoints only, per team decision. Endpoints that write to the production "
         "Supabase database (e.g. submitting a rating, creating a peer group) were excluded from live execution to "
         "avoid polluting real data; see 'Environment & data context' below."),
    ]
    for label, val in meta:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        c = ws.cell(row=r, column=2, value=val)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 14 * (len(val) // 95 + 1))
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Result Summary").font = Font(bold=True, size=12, color=HEADER_FILL)
    r += 1
    total = sum(counts.values())
    headers = ["Result", "Count", "% of tests"]
    for i, h in enumerate(headers):
        ws.cell(row=r, column=1 + i, value=h)
    style_header(ws, r, 3)
    r += 1
    for res in ["PASS", "PARTIAL", "FAIL", "BLOCKED"]:
        n = counts.get(res, 0)
        ws.cell(row=r, column=1, value=res).fill = PatternFill("solid", fgColor=RESULT_COLORS[res])
        ws.cell(row=r, column=1).font = Font(bold=True, color=RESULT_FONT[res])
        ws.cell(row=r, column=2, value=n)
        ws.cell(row=r, column=3, value=f"{(n/total*100):.0f}%" if total else "0%")
        for c in range(1, 4):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value=total).font = Font(bold=True)
    r += 2

    ws.cell(row=r, column=1, value="Environment & Data Context").font = Font(bold=True, size=12, color=HEADER_FILL)
    r += 1
    for note in env_notes:
        c = ws.cell(row=r, column=1, value="•")
        c2 = ws.cell(row=r, column=2, value=note)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 14 * (len(note) // 95 + 1))
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Key Findings (ranked by severity)").font = Font(bold=True, size=12, color=HEADER_FILL)
    r += 1
    for sev, text in key_findings:
        cell = ws.cell(row=r, column=1, value=sev)
        cell.fill = PatternFill("solid", fgColor=SEV_COLORS.get(sev, "FFFFFF"))
        cell.font = Font(bold=True, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="top")
        c2 = ws.cell(row=r, column=2, value=text)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = max(15, 14 * (len(text) // 95 + 1))
        for c in range(1, 3):
            ws.cell(row=r, column=c).border = BORDER
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Overall Assessment").font = Font(bold=True, size=12, color=HEADER_FILL)
    r += 1
    c2 = ws.cell(row=r, column=2, value=narrative)
    c2.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(row=r, column=1, value=narrative)
    ws.row_dimensions[r].height = 130
    r += 2

    ws.cell(row=r, column=1, value="Sign-off").font = Font(bold=True, size=12, color=HEADER_FILL)
    r += 1
    for label in ["Reviewed by (student/team member):", "Expert / Supervisor sign-off:", "Date:", "Comments:"]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
        ws.cell(row=r, column=2, value="").border = Border(bottom=Side(style="thin", color="9CA3AF"))
        r += 1


def write_results_sheet(wb, rows):
    ws = wb.create_sheet("Test Results")
    ws.sheet_view.showGridLines = False
    headers = ["Test ID", "Feature", "Endpoint", "Test Input (summary)", "Validation Criteria",
               "Actual Output (evidence)", "Result", "Severity", "Evaluator Notes / Root Cause"]
    widths = [9, 22, 30, 34, 30, 42, 10, 10, 50]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        ws.cell(row=1, column=i, value=h)
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    r = 2
    for row in rows:
        vals = [row["id"], row["feature"], row["endpoint"], row["input"], row["criteria"],
                row["actual"], row["result"], row.get("severity", ""), row["notes"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
        res_cell = ws.cell(row=r, column=7)
        res_cell.fill = PatternFill("solid", fgColor=RESULT_COLORS[row["result"]])
        res_cell.font = Font(bold=True, color=RESULT_FONT[row["result"]])
        res_cell.alignment = Alignment(horizontal="center", vertical="top")
        sev = row.get("severity", "")
        sev_cell = ws.cell(row=r, column=8)
        sev_cell.fill = PatternFill("solid", fgColor=SEV_COLORS.get(sev, "FFFFFF"))
        sev_cell.alignment = Alignment(horizontal="center", vertical="top")
        # Row height heuristic based on longest wrapped cell
        longest = max(len(str(row["input"])), len(str(row["actual"])), len(str(row["notes"])), len(str(row["criteria"])))
        ws.row_dimensions[r].height = max(30, min(220, 13 * (longest // 40 + 1)))
        r += 1

    ws.auto_filter.ref = f"A1:I{r-1}"


def build_workbook(path, module_no, module_name, owner, service, rows, env_notes, key_findings, narrative):
    counts = {"PASS": 0, "PARTIAL": 0, "FAIL": 0, "BLOCKED": 0}
    for row in rows:
        counts[row["result"]] += 1
    wb = Workbook()
    write_summary_sheet(wb, module_no, module_name, owner, service, counts, env_notes, narrative, key_findings)
    write_results_sheet(wb, rows)
    wb.save(path)
    print(f"Wrote {path}  ({sum(counts.values())} test cases: {counts})")
